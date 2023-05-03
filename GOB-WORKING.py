# With GUI!!!

import pandas as pd
import requests
import json
import argparse
import time
import tkinter as tk
from tkinter import messagebox
import tkinter.scrolledtext as tkst
import sys
from tkinter import messagebox
from io import StringIO
import os
import shutil
import time
import threading
from tkinter import ttk
from tkinter import filedialog
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from tqdm import tqdm
import datetime
import random
import openpyxl.styles
from termcolor import colored
import tkinter as tk


class TerminalWindow:
    def __init__(self, parent):
        self.parent = parent
        self.text_widget = tk.Text(parent)
        self.text_widget.pack(fill="both", expand=True)
        self.stdout_backup = sys.stdout
        self.stderr_backup = sys.stderr
        sys.stdout = self
        sys.stderr = self

    def write(self, message):
        self.text_widget.insert("end", message)
        self.text_widget.see("end")  # scroll to the bottom
        self.flush()

    def flush(self):
        pass

    def close(self):
        sys.stdout = self.stdout_backup
        sys.stderr = self.stderr_backup


class RawDataTab:
    def __init__(self, parent):
        self.parent = parent
        self.frame = tk.Frame(parent)
        self.frame.pack(fill="both", expand=True)

        self.refresh_button = tk.Button(self.frame, text="Refresh", command=self.refresh)
        self.refresh_button.pack()

        self.listbox = tk.Listbox(self.frame)
        self.listbox.pack(fill="both", expand=True)

        self.refresh()

    def refresh(self):
        self.listbox.delete(0, tk.END)
        if not os.path.exists("RawData"):
            self.listbox.insert(tk.END, "RawData folder does not exist.")
            return
        files = os.listdir("RawData")
        if not files:
            self.listbox.insert(tk.END, "RawData folder is empty.")
            return
        for file in files:
            self.listbox.insert(tk.END, file)

class RawDataExcelTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.choose_folder_button = tk.Button(self, text="Choose Created RawData Folder", command=self.choose_folder, width=22, height=2, bg="purple")
        self.choose_folder_button.pack(pady=10, anchor="center")
        self.choose_custom_folder_button = tk.Button(self, text="Choose Custom RawData Folder", command=self.choose_custom_folder, width=25, height=2, bg="purple")
        self.choose_custom_folder_button.pack(pady=10, anchor="center")
        self.create_excel_button = tk.Button(self, text="Create Excel Files", command=self.create_excel_files, width=20, height=2, bg="purple")
        self.create_excel_button.pack(pady=10, anchor="center")
        self.folder_path = ""

    def choose_folder(self):
        # Check if the RawData folder exists in the current working directory
        default_path = os.path.join(os.getcwd(), "RawData")
        if os.path.exists(default_path):
            # Use the default path
            self.folder_path = default_path
            print(f"Selected folder: {self.folder_path}")
            return

        # Display a message box to inform the user that the RawData folder does not exist in the current working directory
        messagebox.showerror("Error", "The RawData folder does not exist in the current working directory.")

        # Set the folder path to None
        self.folder_path = None

        # Display the selected folder path
        print(f"Selected folder: {self.folder_path}")

    def choose_custom_folder(self):
        self.folder_path = filedialog.askdirectory(initialdir=os.getcwd(), title="Select Custom RawData Folder")

        # Check if the selected folder is empty or not named "RawData"
        while not self.folder_path or os.path.basename(self.folder_path) != "RawData":
            # Display a message box to ask the user if they want to cancel their selection
            answer = messagebox.askyesno("Error",
                                         "Please select a valid custom folder. Do you want to cancel your selection?")
            if answer:
                # The user wants to cancel their selection, so return None
                return None

            # The user wants to choose a different folder, so display the file dialog again
            self.folder_path = filedialog.askdirectory(initialdir=os.getcwd(), title="Select Custom RawData Folder")

        # Display the selected folder path
        print(f"Selected custom folder: {self.folder_path}")

    def create_excel_files(self):
        # Check if the RawData folder has been selected
        if not self.folder_path or os.path.basename(self.folder_path) != "RawData":
            messagebox.showerror("Error", "Please select a RawData folder before creating the Excel file.")
            return

        # Step 1: Removing all worksheets except for MAIN
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "MAIN"
        for sheet in workbook.sheetnames[1:]:
            workbook.remove(workbook[sheet])

        # Step 2: Importing all CSVs from the RawData folder
        csv_path = "RawData"  # Update this with the folder path containing CSV files
        if not os.path.exists(csv_path):
            print(colored(
                f"Error: RawData folder not found. Please place the RawData folder where the S1-HealthCheck tool is located.",
                "red"))
            exit()

        csv_files = [os.path.join(csv_path, f) for f in os.listdir(csv_path) if f.endswith(".csv")]

        if len(csv_files) == 0:
            print(colored(f"Error: RawData folder is empty. Please add CSV files to the RawData folder.", "red"))
            exit()
        else:
            print("\033[1;32mSuccess:\033[0m RawData folder found! tool is working...")

        # Step 3: Naming the sheet based on the string left to right until the first dash
        with tqdm(total=len(csv_files), desc="Processing CSV files") as pbar:
            for file in csv_files:
                sheet_name = os.path.basename(file).split("-")[0]
                df = pd.read_csv(file)
                worksheet = workbook.create_sheet(sheet_name)
                worksheet.append(df.columns.tolist())
                for row in df.itertuples(index=False):
                    worksheet.append(row)
                pbar.update(1)

        # Step 7: Adding additional sheets, copying over columns, hiding source, splitting columns
        with tqdm(total=1, desc="Processing policies sheet") as pbar:
            if "policies" in workbook.sheetnames:
                policies_worksheet = workbook["policies"]
                if policies_worksheet:
                    agent_ui_worksheet = workbook.create_sheet("agentUi")
                    agent_ui_worksheet.column_dimensions.group(start='A', end='A', hidden=True)
                    agent_ui_worksheet.append(["AGENT", "SCOPE"])
                    for row in policies_worksheet.iter_rows(min_row=2, values_only=True):
                        if row[46]:  # Column 47 holds the SCOPE
                            scope = row[46].split("-")[-1]
                        agent_ui_worksheet.append([row[21], scope])  # Column 22 holds the AGENT
                policies_worksheet.sheet_state = "hidden"
            else:
                print(colored(
                    f"Error: policies worksheet not found. Please make sure the policies worksheet CSV is included in the RawData folder.",
                    "red"))
                exit()

        # Step 7: Adding additional sheets, copying over columns, hiding source, splitting columns
        with tqdm(total=1, desc="Processing MY_policies sheet") as pbar:
            if "My_Policies" not in workbook.sheetnames:
                workbook.create_sheet("My_Policies")
            policies_worksheet = workbook["policies"]
            my_policies_worksheet = workbook["My_Policies"]

            # Copying over columns
            for row in policies_worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    my_policies_worksheet[cell.coordinate].value = cell.value

            # Copying cell information
            for row in policies_worksheet.iter_rows(min_row=2, values_only=True):
                my_policies_worksheet.append(row)

            # Hiding columns
            my_policies_worksheet.column_dimensions['D'].hidden = True  # Hide agentUi column D
            my_policies_worksheet.column_dimensions['O'].hidden = True  # Hide dvAttributesPerEventType column O
            my_policies_worksheet.column_dimensions['P'].hidden = True  # Hide engines column P
            my_policies_worksheet.column_dimensions['S'].hidden = True  # Hide iocAttributes column S
            my_policies_worksheet.column_dimensions['AA'].hidden = True  # Hide remoteScriptOrchestration column AA

            # Set the fill colors
            light_green = openpyxl.styles.PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            dark_green = openpyxl.styles.Font(color="006400")
            light_red = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            dark_red = openpyxl.styles.Font(color="FF0000")

            # Loop through the cells in the "My_Policies" sheet
            num_rows = my_policies_worksheet.max_row
            with tqdm(total=num_rows) as pbar:
                for row in my_policies_worksheet.iter_rows():
                    for cell in row:
                        # Convert cell value to lowercase for case-insensitive match
                        cell_value = str(cell.value).lower()
                        # Check if the cell value contains "true"
                        if "true" in cell_value:
                            # Apply the light green fill color with dark green background
                            cell.fill = light_green
                            cell.font = dark_green  # Set font color to Dark Green
                        # Check if the cell value contains "false"
                        elif "false" in cell_value:
                            # Apply the light red fill color with dark red background
                            cell.fill = light_red
                            cell.font = dark_red  # Set font color to Dark Red
                    pbar.update(1)

        # Select worksheet
        worksheet = workbook["My_Policies"]

        # Add new sheet policies_DV and copy columns 47 and 19
        policies_dv_sheet = workbook.create_sheet("policies_DV")

        # Copy and transform data from 'policies' sheet to 'policies_DV' sheet
        with tqdm(total=1, desc="'policies' sheet to 'policies_dv' sheet") as pbar:
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                policies_dv_sheet.append([row[46], row[18]])
            pbar.update(1)

        # Hide column 19
        policies_dv_sheet.column_dimensions[get_column_letter(2)].hidden = True

        # Step 1: Split text in column 2 using delimiter ","
        destination_col = []

        for cell in tqdm(policies_dv_sheet["B"], desc="Splitting text..."):
            if isinstance(cell.value, str):
                destination_str = cell.value.strip("{}")  # Remove leading and trailing brackets
                destination_list = destination_str.split(",")
                if len(destination_list) > 0:
                    if destination_list[-1].strip() == "}":
                        destination_list.pop()  # Remove last string from list
                destination_col.append(destination_list)
            else:
                destination_col.append([cell.value])

        # Step 2: Overwrite column 2 with the split text
        for idx, values in enumerate(destination_col):
            for j, value in enumerate(values):
                policies_dv_sheet.cell(row=idx + 1, column=j + 2, value=value)

        # Add conditional formatting for policies_DV
        for row in policies_dv_sheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    if cell.value.strip().endswith("True"):
                        cell.font = openpyxl.styles.Font(color="006400")  # dark green
                        cell.fill = openpyxl.styles.PatternFill(start_color="98FB98", end_color="98FB98",
                                                                fill_type="solid")  # light green
                    elif cell.value.strip().endswith("False"):
                        cell.font = openpyxl.styles.Font(color="FF0000")  # red
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                fill_type="solid")  # lighter red

        # Hide source sheet and create policies_agentUi sheet
        if "policies" in workbook.sheetnames:
            policies_worksheet = workbook["policies"]
            policies_worksheet.column_dimensions[get_column_letter(4)].hidden = False
            policies_agent_ui_worksheet = workbook.create_sheet(title="policies_agentUi",
                                                                index=workbook.index(policies_worksheet) + 1)
            policies_agent_ui_worksheet.column_dimensions.group('A', hidden=False)
            policies_agent_ui_worksheet.append(["SCOPE", "AGENT"])
            for row in policies_worksheet.iter_rows(min_row=2, values_only=True):
                if row[46]:
                    scope = str(row[46]).split("-")[-1]
                    policies_agent_ui_worksheet.append([scope, row[21]])
            policies_agent_ui_worksheet.sheet_state = "hidden"

        # Step 9: Naming the columns
        with tqdm(total=1, desc="Naming the columns") as pbar:
            policies_dv_sheet["B1"] = "Automatically install Deep Visibility browser extensions"
            policies_dv_sheet["C1"] = "Behavioral Indicators"
            policies_dv_sheet["D1"] = "Command Scripts"
            policies_dv_sheet["E1"] = "Cross Process"
            policies_dv_sheet["F1"] = "Data Masking"
            policies_dv_sheet["G1"] = "DLL Module Load"
            policies_dv_sheet["H1"] = "DNS"
            policies_dv_sheet["I1"] = "Full Disk Scan"
            policies_dv_sheet["J1"] = "File"
            policies_dv_sheet["K1"] = "IP"
            policies_dv_sheet["L1"] = "Login"
            policies_dv_sheet["M1"] = "N/A - Not configurable via Policy -- May change in the future"
            policies_dv_sheet["N1"] = "N/A - Not configurable via Policy -- May change in the future"
            policies_dv_sheet["O1"] = "Process"
            policies_dv_sheet["P1"] = "Registry"
            policies_dv_sheet["Q1"] = "Scheduled Tasks"
            policies_dv_sheet["R1"] = "N/A - Not configurable via Policy -- May change in the future"
            policies_dv_sheet["S1"] = "URL"
            pbar.update(1)

        # Step 10: Add color to the first row on each sheet
        with tqdm(total=1, desc="Add color to the first row on each sheet") as pbar:
            fill = PatternFill(start_color='4916ad', end_color='4916ad', fill_type='solid')
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for cell in worksheet["1"]:
                    cell.fill = fill
            pbar.update(1)

        # Create new worksheet policies_engines from policies worksheet
        worksheet = workbook["policies"]
        policies_engines_worksheet = workbook.create_sheet("policies_engines")

        # Copy and transform data from 'policies' sheet to 'policies_engines' sheet
        with tqdm(total=1, desc="'policies' sheet to 'policies_engine' sheet") as pbar:
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                policies_engines_worksheet.append([row[46], row[19]])
            pbar.update(1)

        # Step 1: Split text in column P using delimiter ","
        destination_col = []
        for cell in tqdm(worksheet["P"], desc="Splitting text in column P"):
            if isinstance(cell.value, str):
                destination_str = cell.value.strip("{}")  # Remove leading and trailing brackets
                destination_list = destination_str.split(",")
                if len(destination_list) > 0:
                    if destination_list[-1].strip() == "}":
                        destination_list.pop()  # Remove last string from list
                destination_col.append(destination_list)
            else:
                destination_col.append([cell.value])

        # Step 2: Overwrite column B in policies_engines_worksheet with the split text
        for idx, values in enumerate(destination_col):
            for j, value in enumerate(values):
                policies_engines_worksheet.cell(row=idx + 1, column=j + 2, value=value)

        # Step 3: Add conditional formatting to policies_engines_worksheet
        for row in policies_engines_worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    if cell.value.strip().endswith("'on'"):
                        cell.font = openpyxl.styles.Font(color="006400")  # dark green
                        cell.fill = openpyxl.styles.PatternFill(start_color="98FB98", end_color="98FB98",
                                                                fill_type="solid")  # light green
                    elif cell.value.strip().endswith("'off'"):
                        cell.font = openpyxl.styles.Font(color="FF0000")  # red
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                fill_type="solid")  # lighter red

        # Step 12: Naming the columns
        with tqdm(total=1, desc="Naming the policies_engines columns") as pbar:
            policies_engines_worksheet["B1"].value = "Application Control - Containers only"
            policies_engines_worksheet["C1"].value = "Documents and Scripts"
            policies_engines_worksheet["D1"].value = "Behavioral AI - Executables"
            policies_engines_worksheet["E1"].value = "Anti Exploitation / Fileless"
            policies_engines_worksheet["F1"].value = "Lateral Movement"
            policies_engines_worksheet["G1"].value = "Detect Interactive Threat"
            policies_engines_worksheet["H1"].value = "Static AI"
            policies_engines_worksheet["I1"].value = "Static AI Suspicious"
            policies_engines_worksheet["J1"].value = "Potentially Unwanted Applications"
            policies_engines_worksheet["K1"].value = "Remote Shell Engine (cannot be toggled)"
            policies_engines_worksheet["L1"].value = "Reputation (cannot be toggled)"
            policies_engines_worksheet.auto_filter.ref = "A1:L1"
            pbar.update(1)

        # Step 13: Add color to the first row on each sheet:
        with tqdm(total=1, desc="Add color to the first row on each sheet") as pbar:
            for sheet in workbook:
                for cell in sheet[1]:
                    if cell.value:
                        cell.fill = PatternFill(start_color='4916ad', end_color='4916ad', fill_type='solid')
            pbar.update(1)

        # Step 6: Adding autofilter and autofit to sheets after creation
        with tqdm(total=len(workbook.sheetnames), desc="Adding autofilter and autofit workbook") as pbar:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                worksheet.auto_filter.ref = worksheet.dimensions
                for column in worksheet.columns:
                    max_length = 0
                    column = list(column)
                    column_name = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_name].width = adjusted_width
                pbar.update(1)

        # remoteScriptOrchestration - Sheet will be hidden
        worksheet_rso = workbook.create_sheet("policies_RSO")
        workbook.move_sheet(worksheet_rso, workbook.index(workbook['MAIN']) + 1)
        worksheet_policies = workbook['policies']
        for i in tqdm(range(1, worksheet_policies.max_column + 1), desc="Copying headers"):
            worksheet_rso.cell(row=1, column=i).value = worksheet_policies.cell(row=1, column=i + 46).value
        for i in tqdm(range(2, worksheet_policies.max_row + 1), desc="Copying data"):
            for j in range(1, 3):
                worksheet_rso.cell(row=i, column=j).value = worksheet_policies.cell(row=i, column=j + 26).value
        worksheet_policies.column_dimensions[worksheet_policies.cell(1, 27).column_letter].hidden = True
        col_B = worksheet_rso['B']
        for cell in tqdm(col_B, desc="Replacing commas with periods"):
            if isinstance(cell.value, str):
                cell.value = cell.value.replace(',', '.', 1)
        col_A, col_B = list(worksheet_rso.columns)[0], list(worksheet_rso.columns)[1]
        for col in [col_A, col_B]:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet_rso.column_dimensions[column].width = adjusted_width
        worksheet_rso.sheet_state = 'hidden'

        # Step 4: Add color to the first row on each sheet and hiding column A
        # define sheets to include
        sheets_to_include = ["agent_counts", "policies_DV", "policies_engines"]

        # iterate through sheets and apply formatting
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # set color for first row
            fill = openpyxl.styles.PatternFill(start_color='4916ad', end_color='4916ad', fill_type='solid')
            for cell in worksheet[1]:
                cell.fill = fill
                cell.font = openpyxl.styles.Font(color='FFFFFF')

            # hide column A if not included
            if sheet_name not in sheets_to_include:
                worksheet.column_dimensions.group(start='A', end='A', hidden=True)

        # Hiding other sheets that we do not need
        with tqdm(total=3, desc="Hiding unnecessary sheets") as pbar:
            for sheet_name in ["levels_accounts", "levels_sites", "levels", "agentUi", "levels_groups"]:
                try:
                    worksheet = workbook[sheet_name]
                except KeyError:
                    continue
                worksheet.sheet_state = "hidden"
                pbar.update(1)

        # Moving the MAIN and MAPPINGS sheet to the front
        with tqdm(total=1, desc="Moving MAIN and MAPPINGS sheets") as pbar:
            try:
                worksheet_mappings = workbook['MAPPINGS']
                workbook.move_sheet(worksheet_mappings, 0)
                pbar.update(1)
            except KeyError:
                pbar.set_postfix_str("Worksheet MAPPINGS does not exist, skipping...")
            worksheet_main = workbook['MAIN']
            workbook.move_sheet(worksheet_main, 0)
            pbar.update(1)

        # Save the workbook
        with tqdm(total=1, desc="Final step! Saving the workbook") as pbar:
            filename = datetime.datetime.now().strftime("Health_Check_%d-%m-%Y_{}.xlsx".format(random.randint(0, 999)))
            workbook.save(filename)
            pbar.update(1)

        print("Your New Excel File is ready:", filename)


class GUIInput:
    def __init__(self):
        # self.token_default = "HOep4dW0ocBDfcS6TP3wau5SkCVS0500KlmEAmEXBX1KB4iTYsI570GjzytZcJL6MwNylKVmI0HLX55s"
        self.token_default = "RMqHGGXQcLoyPnded7mUABH2vq5XaB0Rzh4snFS1IAVJgkFnr99DqwDw4gRV7O2dWeHq5GkEZpBxM81q" #allsupport3
        self.domain_default = "usea1-support3.sentinelone.net"
        self.user_default = ""

        self.root = tk.Tk()
        self.root.title("S1-GOB: Guided On Boarding ")
        self.root.iconbitmap("icon.ico")

        # Set application icon for different platforms
        if os.name == 'nt':  # For Windows
            icon_path = "icon.ico"
            self.root.iconbitmap(default=icon_path)
        elif os.name == 'posix':  # For macOS and Linux
            icon_path = "icon.png"
            self.root.iconphoto(True, tk.PhotoImage(file=icon_path))

        # Notebook (tabbed layout)
        self.notebook = ttk.Notebook(self.root, style='My.TNotebook')
        self.notebook.pack(fill='both', expand=True)
        self.notebook.bind("<Button-3>", lambda e: self.popup_menu(e, self.notebook))

        style = ttk.Style()
        style.configure('My.TNotebook', tabposition='n')
        style.configure('My.TNotebook.Tab', padding=[30, 10])

        # API input tab
        api_tab = tk.Frame(self.notebook)
        self.notebook.add(api_tab, text="API Input")

        # Token input
        tk.Label(api_tab, text="Token without the word 'Token'").pack()
        self.token_entry = tk.Entry(api_tab)
        self.token_entry.insert(0, self.token_default)  # set default value
        self.token_entry.pack()

        # Domain input
        tk.Label(api_tab, text="Domain").pack()
        self.domain_entry = tk.Entry(api_tab)
        self.domain_entry.insert(0, self.domain_default)  # set default value
        self.domain_entry.pack()

        # User input
        tk.Label(api_tab, text="User (optional)").pack()
        self.user_entry = tk.Entry(api_tab)
        self.user_entry.insert(0, self.user_default)  # set default value
        self.user_entry.pack()

        # Info button with tooltip
        info_icon = tk.PhotoImage(file="info.png")
        info_button = tk.Button(api_tab, image=info_icon, command=self.show_tooltip)
        info_button.image = info_icon
        info_button.pack()

        # Submit button
        tk.Button(api_tab, text="Submit", command=self.start_script).pack()

        # RawData tab
        rawdata_tab = RawDataTab(self.notebook)
        self.notebook.add(rawdata_tab.frame, text="RawData")

        # RawDataExcel tab
        rawdata_excel_tab = RawDataExcelTab(self.notebook)
        self.notebook.add(rawdata_excel_tab, text="RawDataExcel")

        # Add Terminal tab
        terminal_tab = ttk.Frame(self.root, name="terminal_tab")
        ttk.Label(terminal_tab, text="Terminal").pack()
        self.terminal = TerminalWindow(terminal_tab)
        terminal_tab.pack()

        self.root.mainloop()

    def show_tooltip(self):
        # Set tooltip background and foreground colors
        tooltip_bg = "#6a1b9a"
        tooltip_fg = "#FFFFFF"

        # Set tooltip text with developer information
        tooltip_text = "Use this tool to retrieve information as CSV files and generate Excel reports for health checks. \nEnter your API token, domain, and user email.\n\n*API Token - Enter your console user's API token without the word 'Token'.\n*Domain - Enter XXX.sentinelone.net without 'https://'.\n*User Email - Enter your email address.\n\n\nDeveloped by Ran Jacobi - ranj@sentinelone.com \nVersion 1.2 2023 "
        # Create custom message box with specified colors
        tooltip_box = tk.Toplevel()
        tooltip_box.title("S1-GOB v1.2 Help & About")
        tooltip_box.geometry("600x250")
        tooltip_box.resizable(width=False, height=False)
        tooltip_box.config(bg=tooltip_bg)

        # Create message label with specified colors and text
        tooltip_label = tk.Label(tooltip_box, text=tooltip_text, bg=tooltip_bg, fg=tooltip_fg,
                                 font=("Arial", 13, "bold"), justify="left", padx=10, pady=10)
        tooltip_label.pack()

        # Add picture to the tooltip
        tooltip_image = tk.PhotoImage(file="logo.png")
        tooltip_image_label = tk.Label(tooltip_box, image=tooltip_image, bg=tooltip_bg)
        tooltip_image_label.pack()

        # Set tooltip box to be on top of other windows
        tooltip_box.lift()
        tooltip_box.attributes("-topmost", True)
        tooltip_box.after_idle(tooltip_box.attributes, '-topmost', False)

        # Set tooltip box to be on top of other windows
        tooltip_box.lift()
        tooltip_box.attributes("-topmost", True)
        tooltip_box.after_idle(tooltip_box.attributes, '-topmost', False)

    def start_script(self):
        self.token = self.token_entry.get()
        self.domain = self.domain_entry.get()
        self.user = self.user_entry.get()

        # Validate that required fields are not empty
        if not self.token or not self.domain:
            messagebox.showerror("Error", "Please enter both token and domain fields.")
            return

        # Display message that script is starting to run
        messagebox.showinfo("S1-GOB", "The script is starting to run...")

        # Start the script in a separate thread
        script_thread = threading.Thread(target=self.run_script)
        script_thread.start()

    def run_script(self):
        # Create progress bar
        progress_bar = ttk.Progressbar(self.root, mode='indeterminate')
        progress_bar.pack(fill='x', padx=10, pady=10)
        progress_bar.start()

        # Get token, domain and user
        token = self.token
        domain = self.domain
        user = self.user

        token_header = 'APIToken ' + token
        real_user = user
        customer_endpoint = "https://" + domain

        print(token_header, file=sys.stdout)
        print(real_user, file=sys.stdout)
        print(customer_endpoint, file=sys.stdout)

        # Rest of the code remains the same

        def httpGet(endpoint):
            url = customer_endpoint + endpoint
            headers = {'Authorization': token_header}
            resp = requests.get(url, headers=headers)
            print(resp)
            print(resp.text)
            return resp

        def httpGetPagination(endpoint):
            query_params = "?limit=100"
            nextPage = True
            df_list = []
            FirstRun = True
            sleepcount = 0
            while nextPage:
                print(endpoint + query_params)
                response_json = json.loads(httpGet(endpoint + query_params).text)
                data = response_json['data']
                df_list.append(pd.DataFrame.from_records(data))
                pagination = response_json['pagination']
                if FirstRun:
                    total_left = pagination['totalItems']
                    FirstRun = False
                try:
                    cursor = pagination['nextCursor']
                except:
                    cursor = None
                if cursor:
                    query_params = "?limit=100&cursor=" + cursor
                else:
                    nextPage = False

                total_left -= 100
                if total_left < 0:
                    total_left = 0

                sleepcount += 1
                if sleepcount >= 5:
                    sleepcount = 0
                print("Items Remaining: ", total_left)
            raw_df = pd.concat(df_list)
            print(endpoint + "\n", raw_df)
            return raw_df

        def httpGetPaginationIds(endpoint, level, level_id_df):
            param = "siteIds"

            if level == "accounts":
                param = "accountIds"
            elif level == "sites":
                param = "siteIds"
            elif level == "groups":
                param = "groupIds"

            df_list = []

            print(level)
            print(param)

            for level_id in level_id_df["id"].to_list():
                query_params_base = "?limit=100" + "&" + param + "=" + level_id
                query_params = "?limit=100" + "&" + param + "=" + level_id
                nextPage = True

                FirstRun = True
                sleepcount = 0
                while nextPage:
                    print(endpoint + query_params)
                    response_json = json.loads(httpGet(endpoint + query_params).text)
                    data = response_json['data']
                    dataframe_tmp = pd.DataFrame.from_records(data)
                    dataframe_tmp["level"] = level
                    dataframe_tmp["level_id"] = level_id
                    df_list.append(dataframe_tmp)
                    pagination = response_json['pagination']
                    if FirstRun:
                        total_left = pagination['totalItems']
                        FirstRun = False
                    try:
                        cursor = pagination['nextCursor']
                    except:
                        cursor = None
                    if cursor:
                        query_params = query_params_base + "&cursor=" + cursor
                    else:
                        nextPage = False

                    total_left -= 100
                    if total_left < 0:
                        total_left = 0

                    sleepcount += 1
                    if sleepcount >= 5:
                        sleepcount = 0
                    print("Items Remaining: ", total_left)
            raw_df = pd.concat(df_list)
            try:
                tmp_policy_df = pd.merge(raw_df, level_df, how='left', left_on='level_id', right_on='id')
                print(tmp_policy_df)
                raw_df = tmp_policy_df
            except:
                print("GlobalPolicy")

            print(endpoint + "\n", raw_df)
            return raw_df

        def getGlobalPolicies():
            tmp_global_list = []
            endpoint = "/web/api/v2.1/tenant/policy"
            query_params = "?limit=100"
            response_json = json.loads(httpGet(endpoint + query_params).text)
            data = response_json['data']
            response_fw = httpGet("/web/api/v2.1/firewall-control/configuration")

            if response_fw.status_code == 200:
                response_fwjson = json.loads(response_fw.text)
                fw_data = response_fwjson['data']
                data['fw_enabled'] = fw_data['enabled']
                data['fw_inheritAllFirewallRules'] = fw_data['inheritAllFirewallRules']
                data['fw_inheritedFrom'] = fw_data['inheritedFrom']
                data['fw_inherits'] = fw_data['inherits']
                data['fw_inheritSettings'] = fw_data['inheritSettings']
                data['fw_locationAware'] = fw_data['locationAware']
                # data['fw_reportBlocked'] = fw_data['reportBlocked']
                data['fw_selectedTags'] = fw_data['selectedTags']
            else:
                data['fw_enabled'] = None
                data['fw_inheritAllFirewallRules'] = None
                data['fw_inheritedFrom'] = None
                data['fw_inherits'] = None
                data['fw_inheritSettings'] = None
                data['fw_locationAware'] = None
                # data['fw_reportBlocked'] = None
                data['fw_selectedTags'] = None

            response_dc = httpGet("/web/api/v2.1/device-control/configuration")
            if response_dc.status_code == 200:
                response_dcjson = json.loads(response_dc.text)
                dc_data = response_dcjson['data']
                data['dc_disableBleCommunication'] = dc_data['disableBleCommunication']
                data['dc_disableRfcomm'] = dc_data['disableRfcomm']
                data['dc_disallowAccessPermissionControl'] = dc_data['disallowAccessPermissionControl']
                data['dc_enabled'] = dc_data['enabled']
                data['dc_inheritedFrom'] = dc_data['inheritedFrom']
                data['dc_inherits'] = dc_data['inherits']
                data['dc_reportApproved'] = dc_data['reportApproved']
                data['dc_reportBlocked'] = dc_data['reportBlocked']
                data['dc_reportReadOnly'] = dc_data['reportReadOnly']
            else:
                data['dc_dc_disableBleCommunication'] = None
                data['dc_disableRfcomm'] = None
                data['dc_disallowAccessPermissionControl'] = None
                data['dc_enabled'] = None
                data['dc_inheritedFrom'] = None
                data['dc_inherits'] = None
                data['dc_reportBlocked'] = None
                data['dc_reportReadOnly'] = None

            tmp_global_list.append(data)
            data_df = pd.DataFrame.from_records(tmp_global_list)
            return data_df

        def createLevelsDF():
            id_url = "/web/api/v2.1/accounts"
            account_df = httpGetPagination(id_url)
            id_url = "/web/api/v2.1/sites"
            query_params = "?limit=100"
            nextPage = True
            df_list = []
            while nextPage:
                print(id_url + query_params)
                response_json = json.loads(httpGet(id_url + query_params).text)
                tmp_data = response_json['data']
                data = tmp_data['sites']
                df_list.append(pd.DataFrame.from_records(data))
                pagination = response_json['pagination']
                try:
                    cursor = pagination['nextCursor']
                except:
                    cursor = None
                if cursor:
                    query_params = "?limit=100&cursor=" + cursor
                else:
                    nextPage = False
            raw_df = pd.concat(df_list)
            site_df = raw_df

            id_url = "/web/api/v2.1/groups"
            group_df = httpGetPagination(id_url)
            id_url = "/web/api/v2.1/accounts"
            account_df = httpGetPagination(id_url)

            id_url = "/web/api/v2.1/sites"
            query_params = "?limit=100"
            nextPage = True
            df_list = []
            sleepcount = 0
            while nextPage:
                print(id_url + query_params)
                response_json = json.loads(httpGet(id_url + query_params).text)
                tmp_data = response_json['data']
                data = tmp_data['sites']
                df_list.append(pd.DataFrame.from_records(data))
                pagination = response_json['pagination']
                try:
                    cursor = pagination['nextCursor']
                except:
                    cursor = None

                sleepcount += 1
                if sleepcount >= 5:
                    sleepcount = 0

                if cursor:
                    query_params = "?limit=100&cursor=" + cursor
                else:
                    nextPage = False
            raw_df = pd.concat(df_list)
            site_df = raw_df

            new_account_df = account_df[['id', 'name']].copy()
            site_df = site_df[['id', 'accountId', 'name']].copy()
            group_df = group_df[['id', 'siteId', 'name']].copy()
            print("SiteDF:\n", site_df)
            print("GroupDF:\n", group_df)
            new_site_df = pd.merge(site_df, new_account_df, how='left', left_on='accountId', right_on='id')

            new_site_df['Scope'] = new_site_df["name_y"] + "\\" + new_site_df["name_x"]
            new_site_df = new_site_df.rename(columns={"id_x": "id"})
            del new_site_df['accountId']
            del new_site_df['name_x']
            del new_site_df['id_y']
            del new_site_df['name_y']
            print("PreJoinSite")
            print(new_site_df)

            new_group_df = pd.merge(group_df, new_site_df, how='left', left_on='siteId', right_on='id')

            print("PreNameJoin")
            print(new_group_df)
            new_group_df['Scope_tmp'] = new_group_df['Scope'] + "\\" + new_group_df['name']
            del new_group_df['siteId']
            del new_group_df['name']
            del new_group_df['Scope']
            del new_group_df['id_y']
            new_group_df = new_group_df.rename(columns={"id_x": "id", "Scope_tmp": "Scope"})
            print("NewGroupDF\n", new_group_df)
            print("NewSiteDF\n", new_site_df)
            new_account_df = new_account_df.rename(columns={"name": "Scope"})
            print("NewAccountDF\n", new_account_df)

            return new_account_df, new_site_df, new_group_df

        # Levels = accounts, sites, groups
        def getPolicies(level, level_df):
            if level == "accounts":
                id_url = "/web/api/v2.1/accounts"
                fw_param = "accountIds"
            elif level == "sites":
                id_url = "/web/api/v2.1/sites"
                fw_param = "siteIds"
            elif level == "groups":
                id_url = "/web/api/v2.1/groups"
                fw_param = "groupIds"

            else:
                id_url = "/web/api/v2.1/accounts"

            if level == "global":
                policy_df = getGlobalPolicies()
                policy_df['Scope'] = "Global"
                policy_df['id'] = 0
                policy_df['level_id'] = 0

            else:
                level_df.to_csv("levels-" + level + "-" + domain + ".csv")

                level_id_list = level_df['id'].to_list()
                policy_df_list = []
                iterations_length = len(level_id_list)
                print(iterations_length)
                for level_id in level_id_list:
                    policies_url = id_url + "/" + str(level_id) + "/policy"
                    print(policies_url)
                    response = httpGet(policies_url)
                    if response.status_code == 200:
                        response_json = json.loads(response.text)
                        data = response_json['data']
                        data['level_id'] = level_id
                        fwparam = '?' + fw_param + "=" + level_id
                        response_fw = httpGet("/web/api/v2.1/firewall-control/configuration" + fwparam)
                        if response_fw.status_code == 200:
                            response_fwjson = json.loads(response_fw.text)
                            fw_data = response_fwjson['data']
                            data['fw_enabled'] = fw_data['enabled']
                            data['fw_inheritAllFirewallRules'] = fw_data['inheritAllFirewallRules']
                            data['fw_inheritedFrom'] = fw_data['inheritedFrom']
                            data['fw_inherits'] = fw_data['inherits']
                            data['fw_inheritSettings'] = fw_data['inheritSettings']
                            data['fw_locationAware'] = fw_data['locationAware']
                            # data['fw_reportBlocked'] = fw_data['reportBlocked']
                            data['fw_selectedTags'] = fw_data['selectedTags']
                        else:
                            data['fw_enabled'] = None
                            data['fw_inheritAllFirewallRules'] = None
                            data['fw_inheritedFrom'] = None
                            data['fw_inherits'] = None
                            data['fw_inheritSettings'] = None
                            data['fw_locationAware'] = None
                            # data['fw_reportBlocked'] = None
                            data['fw_selectedTags'] = None

                        response_dc = httpGet("/web/api/v2.1/device-control/configuration" + fwparam)
                        if response_dc.status_code == 200:
                            response_dcjson = json.loads(response_dc.text)
                            dc_data = response_dcjson['data']
                            data['dc_disableBleCommunication'] = dc_data['disableBleCommunication']
                            data['dc_disableRfcomm'] = dc_data['disableRfcomm']
                            data['dc_disallowAccessPermissionControl'] = dc_data['disallowAccessPermissionControl']
                            data['dc_enabled'] = dc_data['enabled']
                            data['dc_inheritedFrom'] = dc_data['inheritedFrom']
                            data['dc_inherits'] = dc_data['inherits']
                            data['dc_reportApproved'] = dc_data['reportApproved']
                            data['dc_reportBlocked'] = dc_data['reportBlocked']
                            data['dc_reportReadOnly'] = dc_data['reportReadOnly']
                        else:
                            data['dc_dc_disableBleCommunication'] = None
                            data['dc_disableRfcomm'] = None
                            data['dc_disallowAccessPermissionControl'] = None
                            data['dc_enabled'] = None
                            data['dc_inheritedFrom'] = None
                            data['dc_inherits'] = None
                            data['dc_reportBlocked'] = None
                            data['dc_reportReadOnly'] = None

                        policy_df_list.append(data)
                    else:
                        print("ERROR Status Code: " + str(response.status_code))
                        print(policies_url)
                        print(response.text)

                    iterations_length -= 1
                    print(level + " Policies left: " + str(iterations_length))
                policy_df = pd.DataFrame.from_records(policy_df_list)

            policy_df['level'] = level
            try:
                tmp_policy_df = pd.merge(policy_df, level_df, how='left', left_on='level_id', right_on='id')
                print(tmp_policy_df)
                policy_df = tmp_policy_df
            except:
                print("GlobalPolicy")

            print("Policy DF " + level + ": ")
            print(policy_df)
            return policy_df

        def getAllPolicies(level_account_df, level_site_df, level_group_df):
            group_policy_df = getPolicies("groups", level_group_df)
            site_policy_df = getPolicies("sites", level_site_df)
            account_policy_df = getPolicies("accounts", level_account_df)
            global_policy_df = getPolicies("global", None)
            policy_frames = [global_policy_df, account_policy_df, site_policy_df, group_policy_df]

            policy_df = pd.concat(policy_frames)
            del policy_df['level_id']
            print(policy_df)

            return (policy_df)

        def getAllLevels(ep, level_account_df, level_site_df, level_group_df):
            global_df = httpGetPagination(ep)
            global_df['level'] = 'Global'
            account_df = httpGetPaginationIds(ep, "account", level_account_df)
            site_df = httpGetPaginationIds(ep, "site", level_site_df)
            group_df = httpGetPaginationIds(ep, "groups", level_group_df)
            policy_frames = [global_df, account_df, site_df, group_df]

            policy_df = pd.concat(policy_frames)
            del policy_df['level_id']
            print(policy_df)
            return policy_df

        level_account_df, level_site_df, level_group_df = createLevelsDF()

        agents_df = httpGetPagination("/web/api/v2.1/agents")
        policy_df = getAllPolicies(level_account_df, level_site_df, level_group_df)
        users_df = httpGetPagination("/web/api/v2.1/users")
        exclusions_df = getAllLevels("/web/api/v2.1/exclusions", level_account_df, level_site_df, level_group_df)
        print(exclusions_df.columns.values.tolist())
        print(exclusions_df)

        # installed_apps_df = getAllLevels("/web/api/v2.1/installed-applications",level_account_df,level_site_df, level_group_df)
        # print(installed_apps_df.columns.values.tolist())
        # print(installed_apps_df)

        try:
            agent_counts = agents_df.groupby(['agentVersion', 'osType']).size()
        except KeyError:
            print("Unable to group by 'agentVersion' and 'osType'")
            agent_counts = None
        try:
            agent_counts.to_csv("agent_counts-" + domain + ".csv")
        except AttributeError:
            print("Error: 'agent_counts' is None")

        agents_df.to_csv("agents-" + domain + ".csv")
        users_df.to_csv("users-" + domain + ".csv")
        policy_df.to_csv("policies-" + domain + ".csv")
        exclusions_df.to_csv("exclusions-" + domain + ".csv")
        # installed_apps_df.to_csv("installed_apps-"+domain+".csv")

        # Create RawData folder if it does not exist
        if not os.path.exists("RawData"):
            os.makedirs("RawData")

        # Move all CSV files to RawData folder
        for file in os.listdir():
            if file.endswith(".csv"):
                shutil.move(file, os.path.join("RawData", file))

        print("\n\n\n\n\n\n\n\n\n\n\n\n\n\n")
        print("Agents DF:\n", agents_df)
        print("Agents Summary:\n", agent_counts)
        print("Users DF:\n", users_df)
        print("Policy DF:\n", policy_df)
        # print("Installed Apps\n",installed_apps_df)
        print("Exclusions DF:\n", exclusions_df)
        print("The Script is Done! The RawData folder is ready")

        # Destroy progress bar
        progress_bar.stop()
        progress_bar.destroy()

        # Display message box
        messagebox.showinfo("Script Status", "Script finished successfully")

gui_input = GUIInput()
exit()
