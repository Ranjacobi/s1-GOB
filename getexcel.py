import datetime
import os
import random

import openpyxl
import openpyxl.styles
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from termcolor import colored
from tqdm import tqdm


class ExcelProcessor:
    def __init__(self, csv_path):
        self.csv_path = csv_path
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "MAIN"

    def process_csv_files(self):
        # Step 1: Removing all worksheets except for MAIN
        for sheet in self.workbook.sheetnames[1:]:
            self.workbook.remove(self.workbook[sheet])

        # Step 2: Importing all CSVs from the RawData folder
        if not os.path.exists(self.csv_path):
            print(colored(f"Error: {self.csv_path} folder not found. Please place the RawData folder where the S1-HealthCheck tool is located.", "red"))
            exit()

        csv_files = [os.path.join(self.csv_path, f) for f in os.listdir(self.csv_path) if f.endswith(".csv")]

        if len(csv_files) == 0:
            print(colored(f"Error: {self.csv_path} folder is empty. Please add CSV files to the RawData folder.", "red"))
            exit()
        else:
            print("Success: starting to create the converted Excel file...")

        # Step 3: Naming the sheet based on the string left to right until the first dash
        pbar = tqdm(total=len(csv_files), desc="Processing CSV files")
        for file in csv_files:
            sheet_name = os.path.basename(file).split("-")[0]
            df = pd.read_csv(file, low_memory=False)
            worksheet = self.workbook.create_sheet(sheet_name)
            worksheet.append(df.columns.tolist())
            for row in df.itertuples(index=False):
                worksheet.append(row)
            pbar.update(1)
        pbar.close()

        # Step 7: Adding additional sheets, copying over columns, hiding source, splitting columns
        with tqdm(total=1, desc="Processing policies sheet") as pbar:
            if "policies" in self.workbook.sheetnames:
                policies_worksheet = self.workbook["policies"]
                if policies_worksheet:
                    agent_ui_worksheet = self.workbook.create_sheet("agentUi")
                    agent_ui_worksheet.column_dimensions.group(start='A', end='A', hidden=True)
                    agent_ui_worksheet.append(["AGENT", "SCOPE"])
                    for row in policies_worksheet.iter_rows(min_row=2, values_only=True):
                        if row[46]:  # Column 47 holds the SCOPE
                            scope = row[46].split("-")[-1]
                        agent_ui_worksheet.append([row[21], scope])  # Column 22 holds the AGENT
                policies_worksheet.sheet_state = "hidden"
            else:
                print(colored(
                    f"Error: policies worksheet not found. Please make sure the policies worksheet CSV is included in the RawData folder.",
                    "red"))
                exit()

        # Step 7: Adding additional sheets, copying over columns, hiding source, splitting columns
        with tqdm(total=1, desc="Processing MY_policies sheet") as pbar:
            if "My_Policies" not in self.workbook.sheetnames:
                self.workbook.create_sheet("My_Policies")
            policies_worksheet = self.workbook["policies"]
            my_policies_worksheet = self.workbook["My_Policies"]

            # Copying over columns
            for row in policies_worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    my_policies_worksheet[cell.coordinate].value = cell.value

            # Copying cell information
            for row in policies_worksheet.iter_rows(min_row=2, values_only=True):
                my_policies_worksheet.append(row)

            # Hiding columns
            my_policies_worksheet.column_dimensions['D'].hidden = True  # Hide agentUi column D
            my_policies_worksheet.column_dimensions['O'].hidden = True  # Hide dvAttributesPerEventType column O
            my_policies_worksheet.column_dimensions['P'].hidden = True  # Hide engines column P
            my_policies_worksheet.column_dimensions['S'].hidden = True  # Hide iocAttributes column S
            my_policies_worksheet.column_dimensions['AA'].hidden = True  # Hide remoteScriptOrchestration column AA

            # Set the fill colors
            light_green = openpyxl.styles.PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            dark_green = openpyxl.styles.Font(color="006400")
            light_red = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            dark_red = openpyxl.styles.Font(color="FF0000")

            # Loop through the cells in the "My_Policies" sheet
            num_rows = my_policies_worksheet.max_row
            with tqdm(total=num_rows) as pbar:
                for row in my_policies_worksheet.iter_rows():
                    for cell in row:
                        # Convert cell value to lowercase for case-insensitive match
                        cell_value = str(cell.value).lower()
                        # Check if the cell value contains "true"
                        if "true" in cell_value:
                            # Apply the light green fill color with dark green background
                            cell.fill = light_green
                            cell.font = dark_green  # Set font color to Dark Green
                        # Check if the cell value contains "false"
                        elif "false" in cell_value:
                            # Apply the light red fill color with dark red background
                            cell.fill = light_red
                            cell.font = dark_red  # Set font color to Dark Red
                pbar.update(1)

        # Select worksheet
        worksheet = self.workbook["My_Policies"]

        # Add new sheet policies_DV and copy columns 47 and 19
        policies_dv_sheet = self.workbook.create_sheet("policies_DV")

        # Copy and transform data from 'policies' sheet to 'policies_DV' sheet
        with tqdm(total=1, desc="'policies' sheet to 'policies_dv' sheet") as pbar:
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                policies_dv_sheet.append([row[46], row[18]])
            pbar.update(1)

        # Hide column 19
        policies_dv_sheet.column_dimensions[get_column_letter(2)].hidden = True

        # Step 1: Split text in column 2 using delimiter ","
        destination_col = []

        for cell in tqdm(policies_dv_sheet["B"], desc="Splitting text..."):
            if isinstance(cell.value, str):
                destination_str = cell.value.strip("{}")  # Remove leading and trailing brackets
                destination_list = destination_str.split(",")
                if len(destination_list) > 0:
                    if destination_list[-1].strip() == "}":
                        destination_list.pop()  # Remove last string from list
                destination_col.append(destination_list)
            else:
                destination_col.append([cell.value])

        # Step 2: Overwrite column 2 with the split text
        for idx, values in enumerate(destination_col):
            for j, value in enumerate(values):
                policies_dv_sheet.cell(row=idx + 1, column=j + 2, value=value)

        # Add conditional formatting for policies_DV
        for row in policies_dv_sheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    if cell.value.strip().endswith("True"):
                        cell.font = openpyxl.styles.Font(color="006400")  # dark green
                        cell.fill = openpyxl.styles.PatternFill(start_color="98FB98", end_color="98FB98",
                                                                fill_type="solid")  # light green
                    elif cell.value.strip().endswith("False"):
                        cell.font = openpyxl.styles.Font(color="FF0000")  # red
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                fill_type="solid")  # lighter red

        # Hide source sheet and create policies_agentUi sheet
        if "policies" in self.workbook.sheetnames:
            policies_worksheet = self.workbook["policies"]
            policies_worksheet.column_dimensions[get_column_letter(4)].hidden = False
            policies_agent_ui_worksheet = self.workbook.create_sheet(title="policies_agentUi",
                                                                index=self.workbook.index(policies_worksheet) + 1)
            policies_agent_ui_worksheet.column_dimensions.group('A', hidden=False)
            policies_agent_ui_worksheet.append(["SCOPE", "AGENT"])
            for row in policies_worksheet.iter_rows(min_row=2, values_only=True):
                if row[46]:
                    scope = str(row[46]).split("-")[-1]
                    policies_agent_ui_worksheet.append([scope, row[21]])
            policies_agent_ui_worksheet.sheet_state = "hidden"

        # Step 9: Naming the columns
        with tqdm(total=1, desc="Naming the columns") as pbar:
            policies_dv_sheet["B1"] = "Automatically install Deep Visibility browser extensions"
            policies_dv_sheet["C1"] = "Behavioral Indicators"
            policies_dv_sheet["D1"] = "Command Scripts"
            policies_dv_sheet["E1"] = "Cross Process"
            policies_dv_sheet["F1"] = "Data Masking"
            policies_dv_sheet["G1"] = "DLL Module Load"
            policies_dv_sheet["H1"] = "DNS"
            policies_dv_sheet["I1"] = "Full Disk Scan"
            policies_dv_sheet["J1"] = "File"
            policies_dv_sheet["K1"] = "IP"
            policies_dv_sheet["L1"] = "Login"
            policies_dv_sheet["M1"] = "N/A - Not configurable via Policy -- May change in the future"
            policies_dv_sheet["N1"] = "N/A - Not configurable via Policy -- May change in the future"
            policies_dv_sheet["O1"] = "Process"
            policies_dv_sheet["P1"] = "Registry"
            policies_dv_sheet["Q1"] = "Scheduled Tasks"
            policies_dv_sheet["R1"] = "N/A - Not configurable via Policy -- May change in the future"
            policies_dv_sheet["S1"] = "URL"
            pbar.update(1)

        # Step 10: Add color to the first row on each sheet
        with tqdm(total=1, desc="Add color to the first row on each sheet") as pbar:
            fill = PatternFill(start_color='4916ad', end_color='4916ad', fill_type='solid')
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                for cell in worksheet["1"]:
                    cell.fill = fill
            pbar.update(1)

        # Create new worksheet policies_engines from policies worksheet
        worksheet = self.workbook["policies"]
        policies_engines_worksheet = self.workbook.create_sheet("policies_engines")

        # Copy and transform data from 'policies' sheet to 'policies_engines' sheet
        with tqdm(total=1, desc="'policies' sheet to 'policies_engine' sheet") as pbar:
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                policies_engines_worksheet.append([row[46], row[19]])
            pbar.update(1)

        # Step 1: Split text in column P using delimiter ","
        destination_col = []
        for cell in tqdm(worksheet["P"], desc="Splitting text in column P"):
            if isinstance(cell.value, str):
                destination_str = cell.value.strip("{}")  # Remove leading and trailing brackets
                destination_list = destination_str.split(",")
                if len(destination_list) > 0:
                    if destination_list[-1].strip() == "}":
                        destination_list.pop()  # Remove last string from list
                destination_col.append(destination_list)
            else:
                destination_col.append([cell.value])

        # Step 2: Overwrite column B in policies_engines_worksheet with the split text
        for idx, values in enumerate(destination_col):
            for j, value in enumerate(values):
                policies_engines_worksheet.cell(row=idx + 1, column=j + 2, value=value)

        # Step 3: Add conditional formatting to policies_engines_worksheet
        for row in policies_engines_worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    if cell.value.strip().endswith("'on'"):
                        cell.font = openpyxl.styles.Font(color="006400")  # dark green
                        cell.fill = openpyxl.styles.PatternFill(start_color="98FB98", end_color="98FB98",
                                                                fill_type="solid")  # light green
                    elif cell.value.strip().endswith("'off'"):
                        cell.font = openpyxl.styles.Font(color="FF0000")  # red
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                fill_type="solid")  # lighter red

        # Step 12: Naming the columns
        with tqdm(total=1, desc="Naming the policies_engines columns") as pbar:
            policies_engines_worksheet["B1"].value = "Application Control - Containers only"
            policies_engines_worksheet["C1"].value = "Documents and Scripts"
            policies_engines_worksheet["D1"].value = "Behavioral AI - Executables"
            policies_engines_worksheet["E1"].value = "Anti Exploitation / Fileless"
            policies_engines_worksheet["F1"].value = "Lateral Movement"
            policies_engines_worksheet["G1"].value = "Detect Interactive Threat"
            policies_engines_worksheet["H1"].value = "Static AI"
            policies_engines_worksheet["I1"].value = "Static AI Suspicious"
            policies_engines_worksheet["J1"].value = "Potentially Unwanted Applications"
            policies_engines_worksheet["K1"].value = "Remote Shell Engine (cannot be toggled)"
            policies_engines_worksheet["L1"].value = "Reputation (cannot be toggled)"
            policies_engines_worksheet.auto_filter.ref = "A1:L1"
            pbar.update(1)

        # Step 13: Add color to the first row on each sheet:
        with tqdm(total=1, desc="Add color to the first row on each sheet") as pbar:
            for sheet in self.workbook:
                for cell in sheet[1]:
                    if cell.value:
                        cell.fill = PatternFill(start_color='4916ad', end_color='4916ad', fill_type='solid')
            pbar.update(1)

        # Step 6: Adding autofilter and autofit to sheets after creation
        with tqdm(total=len(self.workbook.sheetnames), desc="Adding autofilter and autofit workbook") as pbar:
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                worksheet.auto_filter.ref = worksheet.dimensions
                for column in worksheet.columns:
                    max_length = 0
                    column = list(column)
                    column_name = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_name].width = adjusted_width
                pbar.update(1)

        # remoteScriptOrchestration - Sheet will be hidden
        worksheet_rso = self.workbook.create_sheet("policies_RSO")
        self.workbook.move_sheet(worksheet_rso, self.workbook.index(self.workbook['MAIN']) + 1)
        worksheet_policies = self.workbook['policies']
        for i in tqdm(range(1, worksheet_policies.max_column + 1), desc="Copying headers"):
            worksheet_rso.cell(row=1, column=i).value = worksheet_policies.cell(row=1, column=i + 46).value
        for i in tqdm(range(2, worksheet_policies.max_row + 1), desc="Copying data"):
            for j in range(1, 3):
                worksheet_rso.cell(row=i, column=j).value = worksheet_policies.cell(row=i, column=j + 26).value
        worksheet_policies.column_dimensions[worksheet_policies.cell(1, 27).column_letter].hidden = True
        col_B = worksheet_rso['B']
        for cell in tqdm(col_B, desc="Replacing commas with periods"):
            if isinstance(cell.value, str):
                cell.value = cell.value.replace(',', '.', 1)
        col_A, col_B = list(worksheet_rso.columns)[0], list(worksheet_rso.columns)[1]
        for col in [col_A, col_B]:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet_rso.column_dimensions[column].width = adjusted_width
        worksheet_rso.sheet_state = 'hidden'

        # Step 4: Add color to the first row on each sheet and hiding column A
        # define sheets to include
        sheets_to_include = ["agent_counts", "policies_DV", "policies_engines"]

        # iterate through sheets and apply formatting
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]

            # set color for first row
            fill = openpyxl.styles.PatternFill(start_color='4916ad', end_color='4916ad', fill_type='solid')
            for cell in worksheet[1]:
                cell.fill = fill
                cell.font = openpyxl.styles.Font(color='FFFFFF')

            # hide column A if not included
            if sheet_name not in sheets_to_include:
                worksheet.column_dimensions.group(start='A', end='A', hidden=True)

        # Hiding other sheets that we do not need
        with tqdm(total=3, desc="Hiding unnecessary sheets") as pbar:
            for sheet_name in ["levels_accounts", "levels_sites", "levels", "agentUi", "levels_groups"]:
                try:
                    worksheet = self.workbook[sheet_name]
                except KeyError:
                    continue
                worksheet.sheet_state = "hidden"
                pbar.update(1)

        # Save the workbook
        with tqdm(total=1, desc="Final step! Saving the workbook") as pbar:
            filename = datetime.datetime.now().strftime("Health_Check_%d-%m-%Y_{}.xlsx".format(random.randint(0, 999)))
            self.workbook.save(filename)
            pbar.update(1)

        print("Your New Excel File is ready:", filename)
