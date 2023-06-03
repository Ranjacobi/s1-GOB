from tkinter import messagebox
from openpyxl.styles import PatternFill, Font, Border, Side
import glob
import datetime
import random
import time
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import xlrd
from openpyxl import Workbook
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.worksheet import hyperlink

class ExclusionsProcessor:
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "MAIN"

    def combine_functions(self):
        self.process_csv_files()
        self.call_pivot_by_exclusion()
        self.create_pivot_by_exclusion_mode()
        self.search_keywords_and_create_pivot()
        self.create_exclusions_data()

    def process_csv_files(self):
        # Step 1: Importing all CSVs from the RawData folder
        raw_data_folder = os.path.expanduser("~/S1GOB/RawData")
        if not os.path.exists(raw_data_folder):
            try:
                os.makedirs(raw_data_folder)
                print("Success: Created 'RawData' folder.")
            except OSError:
                print("Error: Failed to create 'RawData' folder. Please check write permissions.")
                return

        if not os.access(raw_data_folder, os.W_OK):
            print("Error: Insufficient permissions to write to 'RawData' folder.")
            return

        csv_files = glob.glob(os.path.join(raw_data_folder, "exclusions-*.csv"))

        total_files = len(csv_files)  # Total number of CSV files
        processed_files = 0  # Counter for processed files

        print("Processing the 'exclusions' CSV file...")
        for csv_file in csv_files:
            # Read data from the CSV file
            with open(csv_file, 'r') as file:
                # Assuming the CSV file has comma-separated values
                lines = file.readlines()
                data = [line.strip().split(',') for line in lines]

            # Write data to the Excel worksheet
            for row in data:
                self.worksheet.append(row)

            processed_files += 1

            # Calculate progress percentage
            progress = (processed_files / total_files) * 100
            print(f"Progress: {progress:.2f}%")  # Print progress percentage

            # Check if a progress update is due
            if processed_files % 10 == 0:
                print("Still processing...")

            # Add a delay to simulate processing time
            time.sleep(0.1)

        print("Processing completed.")

    def pivot_by_exclusion(self, sheet_name, exclusion_column, another_exclusion_column):
        # Create a new sheet for the pivot table
        pivot_sheet = self.workbook.create_sheet(title=sheet_name)

        # Convert the sheet data to a DataFrame
        data = list(self.worksheet.values)
        if not data:
            print(f"No data found in sheet 'MAIN'.")
            return

        columns = data.pop(0)
        df = pd.DataFrame(data, columns=columns)

        # Reset index to handle potential duplicate entries
        df.reset_index(inplace=True)

        # Pivot the data by the exclusion column and count the occurrences of 'scopePath'
        pivot_table = df.groupby([exclusion_column, another_exclusion_column]).size().reset_index(name='count')

        # Sort the pivot table by 'count' column in descending order
        pivot_table = pivot_table.sort_values('count', ascending=False)

        # Write the pivot table to the sheet
        for row in dataframe_to_rows(pivot_table, index=False, header=True):
            pivot_sheet.append(row)

    def call_pivot_by_exclusion(self):
        exclusion_column = "description"
        another_exclusion_column = "scopePath"
        sheet_name = "Pivot by Exclusion"
        self.pivot_by_exclusion(sheet_name, exclusion_column, another_exclusion_column)

    def create_pivot_by_exclusion_mode(self):
        # Create a new sheet for the pivot table
        pivot_sheet = self.workbook.create_sheet(title="Pivot by Exclusion Mode")

        # Convert the sheet data to a DataFrame
        data = list(self.worksheet.values)
        if not data:
            print(f"No data found in sheet 'MAIN'.")
            return

        columns = data.pop(0)
        df = pd.DataFrame(data, columns=columns)

        # Reset index to handle potential duplicate entries
        df.reset_index(inplace=True)

        # Pivot the data by the 'mode' column and count the occurrences of 'scopePath'
        pivot_table = df.groupby(['mode', 'scopePath']).size().reset_index(name='count')

        # Sort the pivot table by 'count' column in descending order
        pivot_table = pivot_table.sort_values('count', ascending=False)

        # Write the pivot table to the sheet
        for row in dataframe_to_rows(pivot_table, index=False, header=True):
            pivot_sheet.append(row)

    def search_keywords_and_create_pivot(self):
        # Create a new sheet for "Not Recommended"
        not_recommended_sheet = self.workbook.create_sheet(title="Not Recommended")

        # Load the "not.xlsx" file from the directory
        not_file_path = os.path.expanduser("~/S1GOB/not.xlsx")
        if not os.path.exists(not_file_path):
            print("Error: File 'not.xlsx' not found.")
            return

        # Load the "not.xlsx" workbook
        not_workbook = pd.read_excel(not_file_path)

        # Get the "MAIN" sheet from the workbook
        main_sheet = self.workbook["MAIN"]

        # Get the column index of the "description" column in the "MAIN" sheet
        description_column_index = None
        header_row = main_sheet[1]  # Assuming the header row is the first row
        for index, cell in enumerate(header_row, start=1):
            if cell.value == "description":
                description_column_index = index
                break

        # If "description" column index is found, proceed with comparison
        if description_column_index is not None:
            # Create a dictionary to store the counts and corresponding cell addresses
            count_dict = {}
            cell_address_dict = {}

            # Iterate over each cell in the "not.xlsx" worksheet
            for row in not_workbook.iterrows():
                for cell_value in row[1].values:
                    if pd.notnull(cell_value):
                        # Count the occurrences of cell value in the "MAIN" sheet
                        count = 0
                        cell_addresses = []
                        for row_idx, row in enumerate(main_sheet.iter_rows(min_row=2, values_only=True), start=2):
                            # Assuming data starts from row 2
                            if row[description_column_index - 1] and cell_value.lower() in row[
                                description_column_index - 1].lower():
                                count += 1
                                cell_address = main_sheet.cell(row=row_idx, column=description_column_index).coordinate
                                cell_addresses.append(cell_address)

                        # Store the count and cell addresses in the dictionaries if count is greater than zero
                        if count > 0:
                            count_dict[cell_value] = count
                            cell_address_dict[cell_value] = cell_addresses

            # Convert count_dict and cell_address_dict to DataFrames
            count_df = pd.DataFrame.from_dict(count_dict, orient='index', columns=['Count'])
            count_df.index.name = 'Keyword'
            count_df.reset_index(inplace=True)

            address_df = pd.DataFrame.from_dict(cell_address_dict, orient='index')
            address_df.index.name = 'Keyword'
            address_df.columns = [f'Cell Address {i + 1}' for i in range(address_df.shape[1])]
            address_df.reset_index(inplace=True)

            # Sort the count DataFrame in descending order by counts
            sorted_count_df = count_df.sort_values(by='Count', ascending=False)

            # Merge count_df and address_df on 'Keyword' column
            merged_df = pd.merge(sorted_count_df, address_df, on='Keyword')

            # Write the merged DataFrame to the "Not Recommended" sheet
            not_recommended_sheet.cell(row=1, column=1, value="Keyword")
            not_recommended_sheet.cell(row=1, column=2, value="Count")

            for i, column_name in enumerate(merged_df.columns[2:], start=3):
                not_recommended_sheet.cell(row=1, column=i, value=f"Found In {i - 2}")

            for index, row in merged_df.iterrows():
                keyword = row['Keyword']
                count = row['Count']
                cell_addresses = row.drop(['Keyword', 'Count']).dropna().values.tolist()

                # Write the keyword and count to the corresponding columns
                not_recommended_sheet.cell(row=index + 2, column=1, value=keyword)
                not_recommended_sheet.cell(row=index + 2, column=2, value=count)

                # Write the cell addresses as hyperlinks to the corresponding columns
                for i, cell_address in enumerate(cell_addresses, start=1):
                    column = get_column_letter(i + 2)
                    cell = not_recommended_sheet.cell(row=index + 2, column=2 + i, value=cell_address)
                    hyperlink = f'#MAIN!{cell_address}'
                    cell.hyperlink = hyperlink

        # Print the contents of the "Not Recommended" sheet for testing
        print("Not Recommended Sheet:")
        for row in not_recommended_sheet.iter_rows(values_only=True):
            print(row)

    def create_exclusions_data(self):
        # Iterate over each sheet in the workbook
        sheet_names = self.workbook.sheetnames
        total_sheets = len(sheet_names)
        processed_sheets = 0

        print("Processing sheets...")
        for sheet_name in sheet_names:
            worksheet = self.workbook[sheet_name]

            # Set color for the first row
            fill = PatternFill(start_color='4916ad', end_color='4916ad', fill_type='solid')
            for cell in worksheet[1]:
                cell.fill = fill
                cell.font = Font(color='FFFFFF')

            # Apply borders to the table
            rows = list(worksheet.rows)
            border_style = Side(border_style='thin', color='000000')
            border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            for row in rows:
                for cell in row:
                    cell.border = border

            # Apply autofilter
            worksheet.auto_filter.ref = worksheet.dimensions

            # Autofit columns
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                column_name = column[0].column_letter
                worksheet.column_dimensions[column_name].width = adjusted_width

            processed_sheets += 1

            # Calculate progress percentage
            progress = (processed_sheets / total_sheets) * 100
            print(f"Progress: {progress:.2f}%")  # Print progress percentage
            print("Still processing...")

        # Save the modified Excel file
        filename = datetime.datetime.now().strftime("Exclusions_%d%m%Y_{}.xlsx".format(random.randint(0, 99)))
        s1gob_folder = os.path.expanduser("~/S1GOB")
        if not os.path.exists(s1gob_folder):
            os.makedirs(s1gob_folder)
        file_path = os.path.join(s1gob_folder, filename)
        self.workbook.save(file_path)

        # Wait for 2 seconds
        time.sleep(2)

        print("\n\nYour New Exclusions Excel File is ready:", file_path)

        # Prompt the user if they want to open the new file
        response = messagebox.askquestion("File Created",
                                          f"Do you want to open the new file?\n\nFile name: {file_path}")
        if response == 'yes':
            # Open the file using the default application
            os.system(f'open "{file_path}"')

        print("Success: Excel Exclusions file saved.")
